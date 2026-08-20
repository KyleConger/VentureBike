from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
section_fill = PatternFill("solid", fgColor="D6DCE4")
hint_font = Font(italic=True, size=9, color="666666")
input_fill = PatternFill("solid", fgColor="FFF2CC")
high_fill = PatternFill("solid", fgColor="C6EFCE")
low_fill = PatternFill("solid", fgColor="FFC7CE")
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
green_font = Font(color="006100")
red_font = Font(color="9C0006")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- Instructions ----------
ws = wb.active
ws.title = "Instructions"
ws["A1"] = "MarketingResearch — Q1 Core Data Workbook"
ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
ws["A2"] = (
    "Paste Workspace survey numbers into the yellow input cells. "
    "Thresholds: ≥110 important | <100 may hurt appeal."
)
ws["A2"].font = hint_font

rows = [
    (4, "Sheet", "What to enter", "Analysis use"),
    (5, "Needs", "Importance scores by need × segment", "Sort ≥110; feature translation; peer miss risk"),
    (6, "Applications", "Use-pattern importance by segment", "Component/service fit; peer miss risk"),
    (7, "Price_WTP", "Max price willing to pay (ideal brand)", "Ceiling for Q2 pricing; margin vs elasticity"),
    (8, "Market_Potential", "12-month unit potential by city × segment", "City attractiveness; crowding vs growth room"),
    (9, "Needs_Ranked", "Live links from Needs sheet", "Filter/sort IMPORTANT flags per segment"),
    (10, "Growth_Room", "Attractiveness / crowding / serveability (1–5)", "Competitor-focused opportunity pockets"),
    (11, "Competitor_CI", "Classmate tracker + research purchase log", "Room for growth vs peers (mainly Q2+)"),
]
for r, a, b, c in rows:
    ws.cell(row=r, column=1, value=a)
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c)
    if r == 4:
        style_header(ws, 4, 3)
    else:
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = thin
            ws.cell(row=r, column=col).alignment = left

ws["A13"] = "Primary research focus: maximize competitor (classmate) insight to find room for growth."
ws["A13"].font = Font(bold=True, color="C00000")
ws["A14"] = (
    "Do not assume 25% of annual potential per quarter early — actual early demand is usually much lower."
)
ws["A15"] = "Hard Q1 finance check (elsewhere): Cash + CD ≥ $300,000 after setup decisions."
ws["A17"] = "Market potential formula: Potential = P(purchase) × units × # potential customers"
ws["A19"] = (
    "Need/application row labels are common Marketplace starters — rename to match your Workspace exactly."
)
ws["A19"].font = hint_font
set_widths(ws, [22, 55, 55])

# ---------- Needs ----------
needs_labels = [
    "Comfortable",
    "Easy to ride / handle",
    "Stylish / aesthetics",
    "Durable",
    "Able to carry things",
    "Can handle rough terrain",
    "Can stop quickly",
    "Precise speed control (turns, hills)",
    "Ability to turn sharply",
    "Light weight",
    "Speed",
    "Aerodynamic",
    "Status symbol / exclusivity",
    "Feel young at heart",
    "Soften impact of rough surfaces",
    "Navigation assistance",
    "Other / custom need 1",
    "Other / custom need 2",
    "Other / custom need 3",
    "Other / custom need 4",
]

ws = wb.create_sheet("Needs")
ws["A1"] = "Customer Needs — Importance by Segment"
ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws["A2"] = (
    "Rename labels to match Workspace. Yellow = inputs. "
    "Conditional formatting: green ≥110, red <100."
)
ws["A2"].font = hint_font

headers = [
    "Need / Benefit",
    "Recreation",
    "Mountain",
    "Speed",
    "Max",
    "Rec flag",
    "Mtn flag",
    "Spd flag",
    "Competitor miss risk (H/M/L)",
    "Notes",
]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

for i, label in enumerate(needs_labels):
    r = 5 + i
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=1).fill = input_fill
    ws.cell(row=r, column=5, value=f'=IF(COUNT(B{r}:D{r})=0,"",MAX(B{r}:D{r}))')
    ws.cell(
        row=r,
        column=6,
        value=f'=IF(B{r}="","",IF(B{r}>=110,"IMPORTANT",IF(B{r}<100,"CAUTION","")))',
    )
    ws.cell(
        row=r,
        column=7,
        value=f'=IF(C{r}="","",IF(C{r}>=110,"IMPORTANT",IF(C{r}<100,"CAUTION","")))',
    )
    ws.cell(
        row=r,
        column=8,
        value=f'=IF(D{r}="","",IF(D{r}>=110,"IMPORTANT",IF(D{r}<100,"CAUTION","")))',
    )
    for c in range(1, 11):
        ws.cell(row=r, column=c).border = thin
        ws.cell(row=r, column=c).alignment = center if c > 1 else left
    for c in (2, 3, 4, 9, 10):
        ws.cell(row=r, column=c).fill = input_fill

last_need = 4 + len(needs_labels)
ws.conditional_formatting.add(
    f"B5:D{last_need}",
    CellIsRule(operator="greaterThanOrEqual", formula=["110"], fill=high_fill, font=green_font),
)
ws.conditional_formatting.add(
    f"B5:D{last_need}",
    CellIsRule(operator="lessThan", formula=["100"], fill=low_fill, font=red_font),
)

ws.cell(row=last_need + 2, column=1, value="Segment averages (non-blank)")
ws.cell(row=last_need + 2, column=2, value=f"=IFERROR(AVERAGE(B5:B{last_need}),\"\")")
ws.cell(row=last_need + 2, column=3, value=f"=IFERROR(AVERAGE(C5:C{last_need}),\"\")")
ws.cell(row=last_need + 2, column=4, value=f"=IFERROR(AVERAGE(D5:D{last_need}),\"\")")
set_widths(ws, [36, 12, 12, 12, 8, 12, 12, 12, 18, 40])

# ---------- Applications ----------
apps = [
    "Short social / family rides",
    "Bike paths / light terrain",
    "Carrying small items / groceries",
    "Commuting",
    "Off-road trails",
    "Hills / downhill / cross-country",
    "Vigorous exercise / sport",
    "Long-distance road rides",
    "Racing / group rides",
    "Solo training",
    "Other / custom app 1",
    "Other / custom app 2",
    "Other / custom app 3",
    "Other / custom app 4",
]

ws = wb.create_sheet("Applications")
ws["A1"] = "Applications / Use Patterns — Importance by Segment"
ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws["A2"] = "≥110 = must design for | <100 = may hurt appeal if over-emphasized"
ws["A2"].font = hint_font
headers = [
    "Application / Use",
    "Recreation",
    "Mountain",
    "Speed",
    "Max",
    "Rec flag",
    "Mtn flag",
    "Spd flag",
    "Component implications",
    "Peer miss risk",
]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

for i, label in enumerate(apps):
    r = 5 + i
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=1).fill = input_fill
    ws.cell(row=r, column=5, value=f'=IF(COUNT(B{r}:D{r})=0,"",MAX(B{r}:D{r}))')
    ws.cell(
        row=r,
        column=6,
        value=f'=IF(B{r}="","",IF(B{r}>=110,"IMPORTANT",IF(B{r}<100,"CAUTION","")))',
    )
    ws.cell(
        row=r,
        column=7,
        value=f'=IF(C{r}="","",IF(C{r}>=110,"IMPORTANT",IF(C{r}<100,"CAUTION","")))',
    )
    ws.cell(
        row=r,
        column=8,
        value=f'=IF(D{r}="","",IF(D{r}>=110,"IMPORTANT",IF(D{r}<100,"CAUTION","")))',
    )
    for c in range(1, 11):
        ws.cell(row=r, column=c).border = thin
        ws.cell(row=r, column=c).alignment = center if c > 1 else left
    for c in (2, 3, 4, 9, 10):
        ws.cell(row=r, column=c).fill = input_fill

last_app = 4 + len(apps)
ws.conditional_formatting.add(
    f"B5:D{last_app}",
    CellIsRule(operator="greaterThanOrEqual", formula=["110"], fill=high_fill, font=green_font),
)
ws.conditional_formatting.add(
    f"B5:D{last_app}",
    CellIsRule(operator="lessThan", formula=["100"], fill=low_fill, font=red_font),
)
set_widths(ws, [34, 12, 12, 12, 8, 12, 12, 12, 36, 14])

# ---------- Price WTP ----------
ws = wb.create_sheet("Price_WTP")
ws["A1"] = "Price Willing to Pay (Ideal Brand Ceiling)"
ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws["A2"] = (
    "Treat as maximum for ideal brand. Above = fewer entrants. "
    "Elasticity unknown until test market."
)
ws["A2"].font = hint_font
headers = [
    "Segment",
    "WTP (max)",
    "vs Rec",
    "vs Mountain",
    "Price sensitivity note",
    "Peer price-war risk (H/M/L)",
    "Margin opportunity note",
]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

for i, seg in enumerate(["Recreation", "Mountain", "Speed"]):
    r = 5 + i
    ws.cell(row=r, column=1, value=seg)
    ws.cell(row=r, column=2).fill = input_fill
    ws.cell(row=r, column=3, value=f'=IF(OR($B$5="",B{r}=""),"",B{r}-$B$5)')
    ws.cell(row=r, column=4, value=f'=IF(OR($B$6="",B{r}=""),"",B{r}-$B$6)')
    for c in range(5, 8):
        ws.cell(row=r, column=c).fill = input_fill
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = thin
        ws.cell(row=r, column=c).alignment = center if c > 1 else left

ws["A9"] = "Optional: paste graph / Workspace notes below"
ws.merge_cells("A10:G12")
ws["A10"].fill = input_fill
ws["A10"].border = thin
ws["A10"].alignment = Alignment(vertical="top", wrap_text=True)
set_widths(ws, [14, 12, 10, 12, 36, 18, 36])

# ---------- Market Potential ----------
ws = wb.create_sheet("Market_Potential")
ws["A1"] = "12-Month Market Potential (Units)"
ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws["A2"] = (
    "Potential = P(purchase) × units × # customers. "
    "Early quarterly demand usually << annual/4."
)
ws["A2"].font = hint_font
headers = [
    "City",
    "Recreation",
    "Mountain",
    "Speed",
    "City Total",
    "Naive Qtr (Total/4)",
    "Likely peer crowding (H/M/L)",
    "Growth-room note",
]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

cities = ["New York", "Amsterdam", "Rio de Janeiro", "Bangalore"]
for i, city in enumerate(cities):
    r = 5 + i
    ws.cell(row=r, column=1, value=city)
    for c in (2, 3, 4, 7, 8):
        ws.cell(row=r, column=c).fill = input_fill
    ws.cell(row=r, column=5, value=f'=IF(COUNT(B{r}:D{r})=0,"",SUM(B{r}:D{r}))')
    ws.cell(row=r, column=6, value=f'=IF(E{r}="","",E{r}/4)')
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = thin
        ws.cell(row=r, column=c).alignment = center if c > 1 else left

r = 9
ws.cell(row=r, column=1, value="Segment Total")
ws.cell(row=r, column=1).font = Font(bold=True)
for col_letter, c in zip(["B", "C", "D", "E", "F"], range(2, 7)):
    ws.cell(
        row=r,
        column=c,
        value=f'=IF(COUNT({col_letter}5:{col_letter}8)=0,"",SUM({col_letter}5:{col_letter}8))',
    )
for cc in range(1, 9):
    ws.cell(row=r, column=cc).border = thin
    ws.cell(row=r, column=cc).fill = section_fill

ws["A11"] = "Segment share of total potential"
ws["B11"] = '=IF(OR(B9="",E9="",E9=0),"",B9/E9)'
ws["C11"] = '=IF(OR(C9="",E9="",E9=0),"",C9/E9)'
ws["D11"] = '=IF(OR(D9="",E9="",E9=0),"",D9/E9)'
for cell in ("B11", "C11", "D11"):
    ws[cell].number_format = "0.0%"

ws["A13"] = "City rank by total potential (after fill):"
for i, n in enumerate(["1.", "2.", "3.", "4."], 14):
    ws.cell(row=i, column=1, value=n)
    ws.cell(row=i, column=2).fill = input_fill
    ws.cell(row=i, column=2).border = thin
set_widths(ws, [18, 12, 12, 12, 12, 16, 18, 40])

# ---------- Needs Ranked ----------
ws = wb.create_sheet("Needs_Ranked")
ws["A1"] = "Needs Snapshot — linked to Needs sheet"
ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws["A2"] = (
    "After entering scores on Needs, filter each block where Flag = IMPORTANT, "
    "or sort Score descending (copy-paste values if needed)."
)
ws["A2"].font = hint_font

blocks = [
    (1, "Recreation priorities", "B", "F"),
    (5, "Mountain priorities", "C", "G"),
    (9, "Speed priorities", "D", "H"),
]
for start_col, title, score_col, flag_col in blocks:
    ws.cell(row=4, column=start_col, value=title)
    ws.cell(row=4, column=start_col).font = Font(bold=True, color="1F4E79")
    for offset, h in enumerate(["Need", "Score", "Flag"]):
        cell = ws.cell(row=5, column=start_col + offset, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
    for i in range(len(needs_labels)):
        r = 6 + i
        src = 5 + i
        ws.cell(row=r, column=start_col, value=f"=Needs!A{src}")
        ws.cell(row=r, column=start_col + 1, value=f"=Needs!{score_col}{src}")
        ws.cell(row=r, column=start_col + 2, value=f"=Needs!{flag_col}{src}")
        for c in range(start_col, start_col + 3):
            ws.cell(row=r, column=c).border = thin

set_widths(ws, [28, 10, 12, 3, 28, 10, 12, 3, 28, 10, 12])

# ---------- Growth Room ----------
ws = wb.create_sheet("Growth_Room")
ws["A1"] = "Growth Room vs Classmates (1–5 scores)"
ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws["A2"] = (
    "Gap index = Attractiveness + Serveability − Crowding "
    "(higher = more room for growth). Yellow = inputs."
)
ws["A2"].font = hint_font
headers = [
    "Segment",
    "City / Channel",
    "Attractiveness (1–5)",
    "Expected crowding (1–5)",
    "Our serveability (1–5)",
    "Gap index",
    "Why peers may leave it",
    "Priority (H/M/L)",
]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

combos = []
for seg in ["Recreation", "Mountain", "Speed"]:
    for loc in ["New York", "Amsterdam", "Rio de Janeiro", "Bangalore", "Web"]:
        combos.append((seg, loc))

for i, (seg, loc) in enumerate(combos):
    r = 5 + i
    ws.cell(row=r, column=1, value=seg)
    ws.cell(row=r, column=2, value=loc)
    for c in (3, 4, 5, 7, 8):
        ws.cell(row=r, column=c).fill = input_fill
    ws.cell(row=r, column=6, value=f'=IF(COUNT(C{r}:E{r})<3,"",C{r}+E{r}-D{r})')
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = thin
        ws.cell(row=r, column=c).alignment = center if c != 7 else left

dv = DataValidation(
    type="whole",
    operator="between",
    formula1="1",
    formula2="5",
    allow_blank=True,
)
dv.error = "Enter 1–5"
dv.errorTitle = "Invalid score"
ws.add_data_validation(dv)
dv.add(f"C5:E{4 + len(combos)}")
set_widths(ws, [14, 16, 16, 16, 16, 12, 40, 12])

# ---------- Competitor CI ----------
ws = wb.create_sheet("Competitor_CI")
ws["A1"] = "Competitor (Classmate) Tracker"
ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws["A2"] = "Q1: enter firm names when known. Metrics mainly from Q2+ results / purchased research."
ws["A2"].font = hint_font
headers = [
    "Firm",
    "Quarter",
    "Segments",
    "Cities / Web",
    "Brand Judgment notes",
    "Ad Judgment notes",
    "Price posture",
    "Share / demand served",
    "Weakness / gap for us",
    "Our response",
]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))
for r in range(5, 25):
    for c in range(1, 11):
        ws.cell(row=r, column=c).fill = input_fill
        ws.cell(row=r, column=c).border = thin
set_widths(ws, [18, 10, 16, 18, 22, 20, 14, 18, 28, 28])

ws["A27"] = "Research purchase log (competitor-first)"
ws["A27"].font = Font(bold=True)
headers2 = ["Quarter", "Report bought", "Cost", "Key competitor insight", "Action"]
for i, h in enumerate(headers2, 1):
    ws.cell(row=28, column=i, value=h)
style_header(ws, 28, 5)
for r in range(29, 36):
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = input_fill
        ws.cell(row=r, column=c).border = thin

out = r"c:\Users\Owner\Desktop\MAN6724 Simulation\quarters\Q1\MarketingResearch.xlsx"
wb.save(out)
print("Wrote", out)
