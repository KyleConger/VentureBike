"""Apply locked Q2 sales + production compensation to Comp_Industry (+ Sales_Ops / Manufacturing).
Run when Q2Data.xlsx is closed."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("quarters/Q2/Q2Data.xlsx")
ws = wb["Comp_Industry"]
green = PatternFill("solid", fgColor="C6EFCE")

# --- Sales force (World Market) ---
ws["B38"] = 19000
ws["B38"].fill = green
ws["B39"] = "Full coverage"
ws["B39"].fill = green
ws["C39"] = 4180
ws["C39"].fill = green
ws["B40"] = 2
ws["B40"].fill = green
ws["C40"] = 1055
ws["C40"].fill = green
ws["B41"] = 1
ws["B41"].fill = green
ws["B43"] = 0.85
ws["B43"].fill = green
ws["C43"] = "LOCKED Q2 — World Market / sales force; productivity 85%"

# --- Production workers (World Market) ---
ws["B48"] = 16800
ws["B48"].fill = green
ws["B49"] = "Expanded coverage"
ws["B49"].fill = green
ws["C49"] = 2520
ws["C49"].fill = green
ws["B50"] = 2
ws["B50"].fill = green
ws["C50"] = 933
ws["C50"].fill = green
ws["B51"] = 3
ws["B51"].fill = green
ws["B53"] = 0.85
ws["B53"].fill = green
ws["C53"] = "LOCKED Q2 — World Market / production; productivity 85% → Manufacturing!B30"

# Scenario summary rows (optional labels if present as blank scenario rows)
# Leave scenario scratch alone.

so = wb["Sales_Ops"]
so["B55"] = 24425
so["B55"].fill = green
so["B56"] = "LOCKED Q2 sales package $24,425 (Comp_Industry)"

mfg = wb["Manufacturing"]
mfg["B30"] = 0.85
mfg["B30"].fill = green
mfg["C30"] = "LOCKED Q2 — sync Comp_Industry production productivity 85%"

wb.save("quarters/Q2/Q2Data.xlsx")
print("Applied sales + production compensation locks to Q2Data.xlsx")
print("  Sales: $24,425 · 85%")
print("  Production: $20,757 · 85% (Manufacturing!B30 synced)")
