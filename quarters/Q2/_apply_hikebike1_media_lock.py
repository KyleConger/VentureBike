"""Lock HikeBike1 web + World Market traditional inserts in Media_Preference.
Run when Q2Data.xlsx is closed."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("quarters/Q2/Q2Data.xlsx")
green = PatternFill("solid", fgColor="C6EFCE")
ws = wb["Media_Preference"]

ws["A32"] = "HikeBike1 — Mountain primary"
ws["B32"] = "Web + World Market magazines"
ws["B32"].fill = green
ws["C32"] = 1000  # web qtr
ws["C32"].fill = green
ws["D32"] = "Biking"
ws["E32"] = 12
ws["E32"].fill = green
ws["F32"] = (
    "LOCKED: web $1,000/qtr; inserts Health1 Biking12 Sport2 News1 = $89,000 "
    "(Leisure/Business/NV=0); ranks brand→trail→gears→tires→adventure→tough→mtns→price"
)
ws["G32"] = "LOCKED 2026-08-27"

# Budget cells
ws["B37"] = 89000
ws["B37"].fill = green
ws["C37"] = "World Market traditional inserts LOCKED"
ws["B38"] = 1000
ws["B38"].fill = green
ws["C38"] = "HikeBike1 web LOCKED"

# Detail block under placement if space — use notes rows
ws["A40"] = "HikeBike1 insert detail (LOCKED)"
ws["A41"] = "Medium"
ws["B41"] = "Inserts"
ws["C41"] = "Cost/insert"
ws["D41"] = "Spend"
for r, name, n, c in [
    (42, "Leisure & Entertainment", 0, 10000),
    (43, "Health & Fitness Magazines", 1, 7000),
    (44, "Biking Magazines", 12, 4500),
    (45, "Sport Magazines", 2, 10000),
    (46, "Business Magazines", 0, 9500),
    (47, "New Venture Magazines", 0, 5500),
    (48, "General News Magazines", 1, 8000),
]:
    ws.cell(r, 1, name)
    ws.cell(r, 2, n)
    ws.cell(r, 2).fill = green
    ws.cell(r, 3, c)
    ws.cell(r, 4, n * c)
ws["A49"] = "TOTAL traditional"
ws["D49"] = 89000
ws["D49"].fill = green
ws["E49"] = "$1,594 of ~$90,594 envelope unspent"

wb.save("quarters/Q2/Q2Data.xlsx")
print("Saved HikeBike1 media locks to Media_Preference ($89k inserts + $1k web)")
