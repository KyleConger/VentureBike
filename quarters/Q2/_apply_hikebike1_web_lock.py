"""Lock HikeBike1 web ad + $1000/qtr in Media_Preference. Run when Q2Data.xlsx is closed."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("quarters/Q2/Q2Data.xlsx")
green = PatternFill("solid", fgColor="C6EFCE")
ws = wb["Media_Preference"]

ws["A32"] = "HikeBike1 — Mountain primary (web)"
ws["B32"] = "Web (World Market)"
ws["B32"].fill = green
ws["C32"] = 1000
ws["C32"].fill = green
ws["F32"] = "LOCKED web $1,000/qtr; ranks: brand→trail→gears→tires→adventure→tough→mtns→price"
ws["G32"] = "LOCKED 2026-08-27"

ws["B38"] = 1000
ws["B38"].fill = green
ws["C38"] = "HikeBike1 web only so far; traditional magazine budget OPEN"

if "Web_SEO" in wb.sheetnames:
    seo = wb["Web_SEO"]
    seo["A1"] = seo["A1"].value  # no-op safety
    # Annotate if a notes cell exists near top
    for r in range(1, 40):
        if seo.cell(r, 1).value and "Status" in str(seo.cell(r, 1).value):
            seo.cell(r, 2, "HikeBike1 → Mountain page candidate LOCKED ad; page deploy OPEN")
            break

wb.save("quarters/Q2/Q2Data.xlsx")
print("Locked HikeBike1 web ad: $1,000/qtr in Media_Preference")
