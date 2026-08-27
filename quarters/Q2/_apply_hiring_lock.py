"""Lock Q2 AMS + Web sales/service hiring. Run when Q2Data.xlsx is closed."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("quarters/Q2/Q2Data.xlsx")
green = PatternFill("solid", fgColor="C6EFCE")
so = wb["Sales_Ops"]

# Training rates (quarterly specialty costs from Workspace)
so["C9"] = 300  # Service
so["C9"].fill = green
so["C10"] = 250  # Recreation
so["C10"].fill = green
so["C11"] = 400  # Mountain
so["C11"].fill = green
so["C12"] = 400  # Speed
so["C12"].fill = green
so["C13"] = 0  # Unassigned training
so["D9"] = "LOCKED Q2 — training only; salary via $24,425 annual package"
so["D11"] = "LOCKED Q2 primary segment"

# Amsterdam: 5 total — 1 Service, 0 Rec, 4 Mountain, 0 Speed
so["C19"] = 5
so["C19"].fill = green
so["D19"] = 1
so["D19"].fill = green
so["E19"] = 0
so["E19"].fill = green
so["F19"] = 4
so["F19"].fill = green
so["G19"] = 0
so["G19"].fill = green

# Web personnel: 3 sales + 2 support
so["C55"] = 3
so["C55"].fill = green
so["D55"] = 2
so["D55"].fill = green
so["B55"] = 24425
so["B55"].fill = green

# World Web specialty mirror: 2 Service (support) + 3 Mountain (sales)
so["D20"] = 2
so["D20"].fill = green
so["E20"] = 0
so["F20"] = 3
so["F20"].fill = green
so["G20"] = 0

# Demand forecast tip already 40/SP; headcount links from Sales_Ops
df = wb["Demand_Forecast"]
df["F7"] = 40
df["F7"].fill = green
df["F8"] = 40
df["F8"].fill = green
df["C17"] = "LOCKED hiring → ~400 total @ 40/SP (AMS 5 + web 5); 100% Hike Bike"

wb.save("quarters/Q2/Q2Data.xlsx")
print("Locked hiring: AMS 5 (1 svc / 4 mtn) · Web 5 (3 sales / 2 support)")
print("  Demand @ 40/SP ≈ 400 units · train AMS $1,900 + web specialties $1,800")
print("  Quarterly salary ≈ 10 × $6,106 = $61,062 (+ training; hire fee TBD)")
