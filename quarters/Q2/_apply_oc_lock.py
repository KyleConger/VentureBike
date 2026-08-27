"""Lock Q2 Manufacturing: OC=20/day, OT=0. Run when Q2Data.xlsx is closed."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("quarters/Q2/Q2Data.xlsx")
green = PatternFill("solid", fgColor="C6EFCE")

m = wb["Manufacturing"]
m["B30"] = 0.85
m["B30"].fill = green
m["C30"] = "LOCKED Q2 — production productivity 85%"
m["B39"] = 20
m["B39"].fill = green
m["D39"] = "LOCKED Q2 — 20/day · 1,300/qtr scheduled; effective ~17/day · ~1,105/qtr @ 85%"
m["B53"] = 0
m["B53"].fill = green
m["C53"] = "LOCKED Q2 — no overtime"

# Cost curve hints at OC 20
m["B43"] = 110
m["B44"] = 48

ps = wb["Production_Sim"]
# Try common cells if they exist
for addr, val, note in [
    ("B39", 20, None),  # may not exist
]:
    pass

# Find OC / OT inputs on Production_Sim
for r in range(1, 80):
    a = ps.cell(r, 1).value
    if a is None:
        continue
    s = str(a).lower()
    if "operating capacity" in s and "changeable" in s or (
        "scheduled" in s and "operating" in s
    ):
        pass

wb.save("quarters/Q2/Q2Data.xlsx")
print("Locked Manufacturing: OC=20/day, OT=0, productivity=85%")
print("  Scheduled 1,300/qtr · effective ~1,105/qtr · max demand ~1,130 without OT")
