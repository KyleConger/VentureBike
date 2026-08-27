"""Lock Q2 city #2 = Rio de Janeiro (opens next quarter). Run when Q2Data.xlsx is closed."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("quarters/Q2/Q2Data.xlsx")
green = PatternFill("solid", fgColor="C6EFCE")

so = wb["Sales_Ops"]
so["A21"] = "Rio de Janeiro (opens next Q)"
so["B21"] = "LOCKED — schedule Q2 / live Q3"
so["B21"].fill = green
so["B83"] = "Yes"
so["B83"].fill = green
so["C83"] = "LOCKED Q2 — full quarter to open; live next quarter"
so["B84"] = "Rio"
so["B84"].fill = green
so["C84"] = "Setup $90,000 · lease $17,000/qtr when live (Stores_Costs)"

sc = wb["Stores_Costs"]
sc["B2"] = "Full quarter to open. LIVE: Amsterdam + World web. City #2 LOCKED: Rio (schedule Q2 → live Q3)."
sc["B7"] = "SCHEDULED (Q2) → OPEN next Q"
sc["B7"].fill = green
sc["F7"] = "LOCKED Q2 city #2 — setup $90k this Q; lease when live"

wb.save("quarters/Q2/Q2Data.xlsx")
print("Locked city #2 = Rio de Janeiro (setup $90,000 this quarter; opens next quarter)")
