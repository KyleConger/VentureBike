"""Apply locked Q2 sales compensation to Comp_Industry + Sales_Ops. Run when Q2Data.xlsx is closed."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

wb = load_workbook("quarters/Q2/Q2Data.xlsx")
ws = wb["Comp_Industry"]
green = PatternFill("solid", fgColor="C6EFCE")

ws["B38"] = 19000
ws["B38"].fill = green
ws["B39"] = "Full coverage"
ws["B39"].fill = green
ws["C39"] = 4180
ws["C39"].fill = green
ws["B40"] = 2
ws["B40"].fill = green
ws["C40"] = 1055
ws["C40"].fill = green
ws["B41"] = 1
ws["B41"].fill = green
ws["B43"] = 0.85
ws["B43"].fill = green
ws["C43"] = "LOCKED Q2 — World Market / sales force; productivity 85%"

for r in range(22, 28):
    if ws.cell(r, 1).value == "Sales people":
        ws.cell(r, 2, 19000)
        ws.cell(r, 3, "Full coverage")
        ws.cell(r, 4, "2 weeks")
        ws.cell(r, 5, "1%")
        ws.cell(r, 6, 24425)
        ws.cell(r, 7, "+$5,115 vs $19,310 ind · prod 85%")
        break

so = wb["Sales_Ops"]
so["B55"] = 24425
so["B55"].fill = green
so["B56"] = "LOCKED Q2 sales package $24,425 (Comp_Industry)"

wb.save("quarters/Q2/Q2Data.xlsx")
print("Applied sales compensation lock to Q2Data.xlsx")
