"""Lock Web_Productivity: Start both at default $3k / $6k budgets. Run when Q2Data.xlsx is closed."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("quarters/Q2/Q2Data.xlsx")
green = PatternFill("solid", fgColor="C6EFCE")
ws = wb["Web_Productivity"]

ws["B7"] = "Y"
ws["B7"].fill = green
ws["F7"] = 3000
ws["F7"].fill = green
ws["B8"] = "Y"
ws["B8"].fill = green
ws["F8"] = 6000
ws["F8"].fill = green

# Setup-due this quarter (first start)
for col, label in [(8, "Setup due this Q?"), (9, "Cash this Q")]:
    pass
# Common layout from earlier: I7/I8 may be cash formulas — set setup flags if present
for addr, val in [("H7", "Y"), ("H8", "Y")]:
    if ws[addr].value is None or str(ws[addr].value).startswith("=") is False:
        # only set if column H looks like setup flag (check header)
        pass

# Check headers for setup flag column
hdr = [ws.cell(6, c).value for c in range(1, 12)]
# If I column is cash formula, leave it
ws["A2"] = (
    "LOCKED Q2: Start BOTH — Toll-free budget $3,000 · Page upgrades budget $6,000. "
    "First-Q cash ≈ setup $3k+$6k + budgets $3k+$6k = $18,000. "
    + (str(ws["A2"].value) if ws["A2"].value and "LOCKED" not in str(ws["A2"].value) else "")
)

wb.save("quarters/Q2/Q2Data.xlsx")
print("Locked Web_Productivity: Start both @ $3k / $6k quarterly")
print("  First-quarter cash (setup+budget) ≈ $18,000")
